#!/usr/bin/env python3
"""
Device ETL: pulls SCCM tables from Databricks, joins them on resource_id
and Netbios_Name0, and produces JSON output for Drata Custom Device Connection.

Users are the authoritative anchor: only devices with a matched user record
are included in the output. Devices without a matching user are counted and logged.

Table configuration:
  - Users are loaded first and anchor all downstream scope.
  - Devices are pulled scoped to user machine names (no LIMIT).
  - TABLE_REGISTRY defines secondary tables (installed_software batched, test workspace).
  - required=True entries exit if env var not set.
  - required=False entries are skipped (null) when env var is empty.
  - Adding a new SCCM table: uncomment one registry line, set the env var.

Usage:
    python scripts/extract_devices.py
    python scripts/extract_devices.py --dry-run   # full pipeline, skip Drata push
    python scripts/extract_devices.py --debug     # print resolved env before running
"""

import argparse
import collections
import json
import os
import sys
import time
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

sys.path.insert(0, str(Path(__file__).parent.parent))

from db.auth import get_client_for, load_env
from db import queries
from db.queries import rows_to_records
from db.transform import transform_all, apply_test_overrides, apply_sandbox_overrides

load_env()


# ---------------------------------------------------------------------------
# Internal column stripping
# ---------------------------------------------------------------------------
STRIP_PREFIXES = ("__",)
_MAX_RETRIES = 3

# Devices whose Netbios_Name0 does not start with one of these prefixes are dropped
# before the merge. Add new prefixes here as needed (case-insensitive).
DEVICE_NAME_PREFIXES = ('NW', 'GI')
_RETRY_DELAYS = (5, 15)  # seconds before attempt 2 and attempt 3
_SW_BATCH_SIZE = 200
_PERSONNEL_CHECK_WORKERS = 5
_PIPELINE_CHUNK_SIZE = 500


def is_internal(col: str) -> bool:
    return any(col.startswith(p) for p in STRIP_PREFIXES)


# ---------------------------------------------------------------------------
# Table registry
# ---------------------------------------------------------------------------
TableSpec = collections.namedtuple(
    'TableSpec', ['label', 'env_var', 'client_key', 'filter_type', 'required', 'batched']
)
# client_key   : 'prod' | 'test'
# filter_type  : 'resource_id' | 'netbios_name'
# required     : True = env var must be set; False = skipped (None) when empty
# batched      : True = use pull_table_batched() (IN-clause chunked to _SW_BATCH_SIZE)

TABLE_REGISTRY = [
    TableSpec('windows_update',     'DATABRICKS_TABLE_WINDOWS_UPDATE',     'test', 'resource_id', True,  False),
    TableSpec('installed_software', 'DATABRICKS_TABLE_INSTALLED_SOFTWARE', 'test', 'resource_id', True,  True),
    # Uncomment when Nationwide confirms table names:
    # TableSpec('bitlocker',       'DATABRICKS_TABLE_BITLOCKER',       'test', 'resource_id', False, False),
    # TableSpec('screensaver',     'DATABRICKS_TABLE_SCREENSAVER',     'test', 'resource_id', False, False),
    # TableSpec('services',        'DATABRICKS_TABLE_SERVICES',        'test', 'resource_id', False, False),
    # TableSpec('network_adapter', 'DATABRICKS_TABLE_NETWORK_ADAPTER', 'test', 'resource_id', False, False),
]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def clean(record: Dict[str, Any]) -> Dict[str, Any]:
    """Strip internal pipeline columns from a record."""
    return {k: v for k, v in record.items() if not is_internal(k)}


def get_resource_id(record: Dict[str, Any]) -> Optional[int]:
    """Resolve resource_id regardless of column name casing."""
    for key in ("resource_id", "ResourceID", "RESOURCEID"):
        if key in record:
            val = record[key]
            try:
                return int(val) if val is not None else None
            except (ValueError, TypeError):
                return None
    return None


def _ids_filter(ids: List[int], column: str = "resource_id") -> str:
    """Build a SQL IN filter for integer IDs. Returns '1=0' if the list is empty."""
    if not ids:
        return "1=0"
    return f"{column} IN ({', '.join(str(int(i)) for i in ids)})"


def _names_filter(names: List[str], column: str = "Netbios_Name0") -> str:
    """Build a SQL IN filter for string names with single-quote escaping."""
    if not names:
        return "1=0"
    escaped = ", ".join("'" + n.replace("'", "''") + "'" for n in names)
    return f"{column} IN ({escaped})"


def pull_table(
    client: Any,
    table: str,
    warehouse_id: str,
    label: str,
    limit: Optional[int] = None,
    filter_sql: Optional[str] = None,
    timeout: int = 300,
) -> List[Dict[str, Any]]:
    """Pull a table and return cleaned records. Exits on failure after retries."""
    parts = [f"SELECT * FROM {table}"]
    if filter_sql:
        parts.append(f"WHERE {filter_sql}")
    if limit is not None:
        parts.append(f"LIMIT {limit}")
    statement = " ".join(parts)
    print(f"  Pulling {label} ({table}) ...")
    last_error: Optional[Exception] = None
    for attempt in range(1, _MAX_RETRIES + 1):
        try:
            result = queries.run_sql(
                client,
                statement=statement,
                warehouse_id=warehouse_id,
                timeout_seconds=timeout,
            )
            records = rows_to_records(result["columns"], result["rows"])
            print(f"  {len(records)} rows retrieved.")
            return [clean(r) for r in records]
        except Exception as e:
            last_error = e
            if attempt < _MAX_RETRIES:
                wait = _RETRY_DELAYS[attempt - 1]
                print(f"  [RETRY {attempt}/{_MAX_RETRIES}] {label} failed, retrying in {wait}s ...")
                time.sleep(wait)
    raw = str(last_error)
    short = raw.split(". Config:")[0].split(". Env:")[0].strip()
    print(f"  [FAIL] {label} (all {_MAX_RETRIES} attempts failed)")
    print(f"         Table     : {table}")
    print(f"         Warehouse : {warehouse_id}")
    print(f"         Error     : {short}")
    sys.exit(1)


def pull_table_batched(
    client: Any,
    table: str,
    warehouse_id: str,
    label: str,
    ids: List[int],
    id_column: str = "resource_id",
    timeout: int = 300,
) -> List[Dict[str, Any]]:
    """Pull a large table by chunking the IN-clause into batches of _SW_BATCH_SIZE."""
    all_records: List[Dict[str, Any]] = []
    total_batches = (len(ids) + _SW_BATCH_SIZE - 1) // _SW_BATCH_SIZE
    for i in range(0, len(ids), _SW_BATCH_SIZE):
        batch = ids[i : i + _SW_BATCH_SIZE]
        batch_num = i // _SW_BATCH_SIZE + 1
        records = pull_table(
            client, table, warehouse_id,
            f"{label} (batch {batch_num}/{total_batches})",
            filter_sql=_ids_filter(batch, column=id_column),
            timeout=timeout,
        )
        all_records.extend(records)
    return all_records


def merge(
    devices: List[Dict[str, Any]],
    windows_update: List[Dict[str, Any]],
    installed_software: List[Dict[str, Any]],
    users: List[Dict[str, Any]],
    bitlocker: Optional[List[Dict[str, Any]]] = None,
    screensaver: Optional[List[Dict[str, Any]]] = None,
    services: Optional[List[Dict[str, Any]]] = None,
    network_adapter: Optional[List[Dict[str, Any]]] = None,
) -> Tuple[List[Dict[str, Any]], int]:
    """
    Inner join: users are the anchor. Only devices with a matched user are included.

    Returns (records, dropped_count) where dropped_count is the number of devices
    that had no matching user entry.
    """
    # Index devices by machine name
    device_index: Dict[str, Dict[str, Any]] = {}
    for dev in devices:
        netbios = dev.get('Netbios_Name0') or dev.get('Name0')
        if netbios:
            device_index[netbios] = dev

    # Index resource_id-keyed tables
    wu_index: Dict[int, Dict[str, Any]] = {}
    for row in windows_update:
        rid = get_resource_id(row)
        if rid is not None:
            wu_index[rid] = {k: v for k, v in row.items() if k not in ("resource_id", "ResourceID")}

    sw_index: Dict[int, List[Dict[str, Any]]] = {}
    for row in installed_software:
        rid = get_resource_id(row)
        if rid is not None:
            entry = {k: v for k, v in row.items() if k not in ("resource_id", "ResourceID")}
            sw_index.setdefault(rid, []).append(entry)

    # Optional tables -- build index only when table was pulled
    bitlocker_index: Optional[Dict[int, Dict[str, Any]]] = None
    if bitlocker is not None:
        bitlocker_index = {}
        for row in bitlocker:
            rid = get_resource_id(row)
            if rid is not None:
                bitlocker_index[rid] = {k: v for k, v in row.items() if k not in ("resource_id", "ResourceID")}

    screensaver_index: Optional[Dict[int, Dict[str, Any]]] = None
    if screensaver is not None:
        screensaver_index = {}
        for row in screensaver:
            rid = get_resource_id(row)
            if rid is not None:
                screensaver_index[rid] = {k: v for k, v in row.items() if k not in ("resource_id", "ResourceID")}

    services_index: Optional[Dict[int, List[Dict[str, Any]]]] = None
    if services is not None:
        services_index = {}
        for row in services:
            rid = get_resource_id(row)
            if rid is not None:
                entry = {k: v for k, v in row.items() if k not in ("resource_id", "ResourceID")}
                services_index.setdefault(rid, []).append(entry)

    network_adapter_index: Optional[Dict[int, Dict[str, Any]]] = None
    if network_adapter is not None:
        network_adapter_index = {}
        for row in network_adapter:
            rid = get_resource_id(row)
            if rid is not None:
                network_adapter_index[rid] = {k: v for k, v in row.items() if k not in ("resource_id", "ResourceID")}

    # User-centric iteration: users anchor the output set
    matched_netbios: set = set()
    output: List[Dict[str, Any]] = []
    for row in users:
        netbios = row.get('Netbios_Name0') or row.get('netbios_name0')
        if not netbios or netbios not in device_index:
            continue
        device = device_index[netbios]
        matched_netbios.add(netbios)
        rid = get_resource_id(device)
        user_fields = {k: v for k, v in row.items() if k not in ('Netbios_Name0', 'netbios_name0')}
        device_fields = {k: v for k, v in device.items() if k not in ("resource_id", "ResourceID", "ResourceType")}
        output.append({
            "resource_id": rid,
            "device": device_fields,
            "windows_update": wu_index.get(rid, {}),
            "installed_software": sw_index.get(rid, []),
            "user": user_fields,
            "bitlocker": bitlocker_index.get(rid) if bitlocker_index is not None else None,
            "screensaver": screensaver_index.get(rid) if screensaver_index is not None else None,
            "services": services_index.get(rid, []) if services_index is not None else None,
            "network_adapter": network_adapter_index.get(rid) if network_adapter_index is not None else None,
        })

    dropped = len(devices) - len(matched_netbios)
    return output, dropped


def write_json(payload: List[Dict[str, Any]], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, default=str)


_LOCAL_USERS_FILE = "SCCM Employees with Devices - Sandbox.xlsx"


def load_users_from_xlsx(path: str, netbios_filter: Optional[set] = None) -> List[Dict[str, Any]]:
    """Load user records from a local xlsx file. When netbios_filter is None, all rows are returned."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = {k: v for k, v in zip(headers, row) if k is not None}
        if netbios_filter is None or record.get('Netbios_Name0') in netbios_filter:
            records.append(record)
    wb.close()
    return records


def default_output_paths(test_mode: bool = False) -> Tuple[Path, Path]:
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    tag = "_test" if test_mode else ""
    return (
        Path("output") / f"devices_{timestamp}{tag}_raw.json",
        Path("output") / f"devices_{timestamp}{tag}_drata.json",
    )


# ---------------------------------------------------------------------------
# Args
# ---------------------------------------------------------------------------

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Pull SCCM tables from Databricks, merge by user, transform to Drata MDM format."
    )
    parser.add_argument(
        "--devices",
        metavar="CATALOG.SCHEMA.TABLE",
        default=os.getenv("DATABRICKS_TABLE_DEVICES", ""),
        help="Fully qualified path to the main device table (prod workspace). Uses DATABRICKS_TABLE_DEVICES.",
    )
    parser.add_argument(
        "--warehouse-prod",
        metavar="WAREHOUSE_ID",
        default=os.getenv("DATABRICKS_WAREHOUSE_ID", ""),
        help="Warehouse ID for the prod devices table. Uses DATABRICKS_WAREHOUSE_ID.",
    )
    parser.add_argument(
        "--warehouse-test",
        metavar="WAREHOUSE_ID",
        default=os.getenv("DATABRICKS_WAREHOUSE_ID_TEST", ""),
        help="Warehouse ID for the test catalog tables. Uses DATABRICKS_WAREHOUSE_ID_TEST.",
    )
    parser.add_argument(
        "--host-prod",
        metavar="URL",
        default=os.getenv("DATABRICKS_HOST_PROD", ""),
        help="Prod workspace URL. Uses DATABRICKS_HOST_PROD.",
    )
    parser.add_argument(
        "--host-test",
        metavar="URL",
        default=os.getenv("DATABRICKS_HOST_TEST", ""),
        help="Test workspace URL. Uses DATABRICKS_HOST_TEST.",
    )
    parser.add_argument(
        "--token-prod",
        metavar="TOKEN",
        default=os.getenv("DATABRICKS_TOKEN_PROD", "") or os.getenv("DATABRICKS_TOKEN", ""),
        help="Token for the prod workspace. Uses DATABRICKS_TOKEN_PROD, falls back to DATABRICKS_TOKEN.",
    )
    parser.add_argument(
        "--token-test",
        metavar="TOKEN",
        default=os.getenv("DATABRICKS_TOKEN_TEST", "") or os.getenv("DATABRICKS_TOKEN", ""),
        help="Token for the test workspace. Uses DATABRICKS_TOKEN_TEST, falls back to DATABRICKS_TOKEN.",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=int(os.getenv("DATABRICKS_LIMIT", "1000")),
        help="Max users to process per run (default: 1000). Devices and secondary tables are scoped to those users.",
    )
    parser.add_argument(
        "--timeout",
        type=int,
        default=int(os.getenv("DATABRICKS_QUERY_TIMEOUT", "300")),
        help="Per-query timeout in seconds (default: 300). Uses DATABRICKS_QUERY_TIMEOUT.",
    )
    parser.add_argument(
        "--output-raw",
        metavar="FILE",
        default="",
        help="Path for the raw merged JSON. Defaults to output/devices_<timestamp>_raw.json.",
    )
    parser.add_argument(
        "--output-drata",
        metavar="FILE",
        default="",
        help="Path for the Drata-formatted JSON. Defaults to output/devices_<timestamp>_drata.json.",
    )
    parser.add_argument(
        "--local-users",
        action="store_true",
        default=False,
        help=(
            f"Load the users table from the local xlsx file ({_LOCAL_USERS_FILE}) "
            "instead of pulling from Databricks. Bypasses DATABRICKS_TABLE_USERS. "
            "Records are scoped to the machine names returned by the devices pull."
        ),
    )
    parser.add_argument(
        "--test-mode",
        action="store_true",
        default=False,
        help=(
            "Push real identities to Drata with all 5 monitoring fields forced to a passing "
            "state. Uses real personnelId/alias/externalId so records land on actual users. "
            "Combine with --limit to control how many are processed. "
            "Intended for verifying the Drata connection end-to-end."
        ),
    )
    parser.add_argument(
        "--sandbox",
        action="store_true",
        default=False,
        help="Replace @nationwide.com with @sandbox.nationwide.com in personnelId before pushing.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        default=False,
        help="Run the full pipeline but skip the Drata API push. Output files are still written.",
    )
    parser.add_argument(
        "--debug",
        action="store_true",
        default=False,
        help="Print full resolved config and env var sources before running.",
    )
    parser.add_argument(
        "--full",
        action="store_true",
        default=False,
        help=(
            "Process all users -- bypasses --limit and runs the full dataset using the "
            "chunked pipeline. Intended for production sync of 25,000+ users."
        ),
    )
    return parser.parse_args()


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    args = parse_args()

    # Validate required workspace/warehouse config upfront
    missing = [name for name, val in [
        ("--devices (or DATABRICKS_TABLE_DEVICES)", args.devices),
        ("--host-prod (or DATABRICKS_HOST_PROD)", args.host_prod),
        ("--host-test (or DATABRICKS_HOST_TEST)", args.host_test),
        ("--warehouse-prod (or DATABRICKS_WAREHOUSE_ID)", args.warehouse_prod),
        ("--warehouse-test (or DATABRICKS_WAREHOUSE_ID_TEST)", args.warehouse_test),
    ] if not val.strip()]

    if missing:
        print("Error: the following required values are not set:")
        for m in missing:
            print(f"  {m}")
        sys.exit(1)

    prod_client = get_client_for(host=args.host_prod, token=args.token_prod)
    test_client = get_client_for(host=args.host_test, token=args.token_test)
    default_raw, default_drata = default_output_paths(test_mode=args.test_mode)
    raw_path = Path(args.output_raw) if args.output_raw else default_raw
    drata_path = Path(args.output_drata) if args.output_drata else default_drata

    print(f"\nProd workspace   : {args.host_prod}")
    print(f"Test workspace   : {args.host_test}")
    print(f"Warehouse (prod) : {args.warehouse_prod}")
    print(f"Warehouse (test) : {args.warehouse_test}")
    print(f"Limit (users)    : {args.limit} users")
    print(f"Query timeout    : {args.timeout}s per table")
    print(f"Output (raw)     : {raw_path}")
    print(f"Output (drata)   : {drata_path}")
    if args.local_users:
        print(f"Users source     : LOCAL FILE ({_LOCAL_USERS_FILE})")
    if args.test_mode:
        print(f"Mode             : TEST MODE (all 5 monitoring fields forced passing)")
    if args.full:
        print(f"Mode             : FULL SYNC (all users, --limit bypassed)")
    if args.dry_run:
        print(f"Mode             : DRY RUN (Drata push skipped)")

    if args.debug:
        databrickscfg = Path.home() / ".databrickscfg"
        print(f"\n-- DEBUG --")
        print(f"DATABRICKS_HOST_PROD             : {os.getenv('DATABRICKS_HOST_PROD', '(not set)')}")
        print(f"DATABRICKS_HOST_TEST             : {os.getenv('DATABRICKS_HOST_TEST', '(not set)')}")
        print(f"DATABRICKS_TOKEN_PROD            : {'(set)' if os.getenv('DATABRICKS_TOKEN_PROD') or os.getenv('DATABRICKS_TOKEN') else '(not set)'}")
        print(f"DATABRICKS_TOKEN_TEST            : {'(set)' if os.getenv('DATABRICKS_TOKEN_TEST') or os.getenv('DATABRICKS_TOKEN') else '(not set)'}")
        print(f"DATABRICKS_WAREHOUSE_ID          : {os.getenv('DATABRICKS_WAREHOUSE_ID', '(not set)')}")
        print(f"DATABRICKS_WAREHOUSE_ID_TEST     : {os.getenv('DATABRICKS_WAREHOUSE_ID_TEST', '(not set)')}")
        print(f"DATABRICKS_TABLE_DEVICES         : {os.getenv('DATABRICKS_TABLE_DEVICES', '(not set)')}")
        for spec in TABLE_REGISTRY:
            val = os.getenv(spec.env_var, '(not set)')
            req = 'required' if spec.required else 'optional'
            print(f"{spec.env_var:<40}: {val}  [{req}]")
        print(f"DATABRICKS_QUERY_TIMEOUT         : {os.getenv('DATABRICKS_QUERY_TIMEOUT', '(not set, using 300)')}")
        print(f"DRATA_API_KEY                    : {'(set)' if os.getenv('DRATA_API_KEY') else '(not set)'}")
        print(f"DRATA_CONNECTION_ID              : {os.getenv('DRATA_CONNECTION_ID', '(not set)')}")
        print(f"~/.databrickscfg exists          : {databrickscfg.exists()}")
        if args.local_users:
            xlsx_exists = Path(_LOCAL_USERS_FILE).exists()
            print(f"LOCAL_USERS_FILE                 : {_LOCAL_USERS_FILE}  ({'found' if xlsx_exists else 'NOT FOUND'})")
        print(f"-- END DEBUG --\n")
    else:
        print()

    api_key = os.getenv("DRATA_API_KEY", "").strip()
    connection_id = os.getenv("DRATA_CONNECTION_ID", "").strip()

    # Step 1: load users first (anchor for all downstream scope)
    if args.local_users:
        if not Path(_LOCAL_USERS_FILE).exists():
            print(f"  [FAIL] Local users file not found: {_LOCAL_USERS_FILE}")
            sys.exit(1)
        print(f"  [LOCAL] Loading users from {_LOCAL_USERS_FILE} ...")
        all_users = load_users_from_xlsx(_LOCAL_USERS_FILE, netbios_filter=None)
        print(f"  {len(all_users)} users loaded.")
    else:
        users_table = os.getenv("DATABRICKS_TABLE_USERS", "").strip()
        if not users_table:
            print("  [FAIL] DATABRICKS_TABLE_USERS is required but not set (or use --local-users)")
            sys.exit(1)
        all_users = pull_table(test_client, users_table, args.warehouse_test, "users", timeout=args.timeout)

    if args.full:
        print(f"  Full sync: processing all {len(all_users)} users.")
    elif args.limit and len(all_users) > args.limit:
        print(f"  Limiting to {args.limit} users (of {len(all_users)} total).")
        all_users = all_users[:args.limit]

    # Step 1b: validate each user against Drata personnel status (parallel)
    if api_key:
        from db.drata_client import DrataClient
        from concurrent.futures import ThreadPoolExecutor, as_completed as _as_completed
        drata_pre = DrataClient(api_key=api_key, connection_id=connection_id)
        upn_key = 'User_Princiipal_Name0'  # double-i typo in source
        _active = {'CURRENT_EMPLOYEE', 'CURRENT_CONTRACTOR'}
        before = len(all_users)
        sandbox_flag = args.sandbox

        def _check_one(u):
            try:
                email = (u.get(upn_key) or u.get('User_Principal_Name0') or '').lower()
                if not email or '@' not in email:
                    return u, None
                lookup = email
                if sandbox_flag and '@nationwide.com' in email:
                    lookup = email.replace('@nationwide.com', '@sandbox.nationwide.com')
                return u, drata_pre.get_person_status(lookup)
            except Exception:
                return u, '__error__'

        filtered = []
        check_errors = 0
        print(f"Checking {before} user(s) against Drata personnel status ({_PERSONNEL_CHECK_WORKERS} workers) ...")
        with ThreadPoolExecutor(max_workers=_PERSONNEL_CHECK_WORKERS) as pool:
            futures = {pool.submit(_check_one, u): u for u in all_users}
            for i, fut in enumerate(_as_completed(futures), 1):
                u, status = fut.result()
                if status == '__error__':
                    check_errors += 1
                elif status in _active:
                    filtered.append(u)
                if i % 500 == 0:
                    print(f"  ... {i}/{before} checked, {len(filtered)} active ...")
        all_users = filtered
        skipped = before - len(all_users) - check_errors
        print(f"  Personnel filter: {len(all_users)} active / {skipped} excluded (former/not found) / {check_errors} API errors (excluded).")
    else:
        print("  [WARN] DRATA_API_KEY not set -- personnel filter skipped, all users will be processed.")

    if not all_users:
        print("  [FAIL] No users remain after personnel filter -- check DRATA_API_KEY or user data")
        sys.exit(1)

    # Steps 2-5: process in chunks to bound memory and Databricks query scope
    chunks = [all_users[i:i + _PIPELINE_CHUNK_SIZE]
              for i in range(0, len(all_users), _PIPELINE_CHUNK_SIZE)]
    print(f"\n{len(all_users)} user(s) in {len(chunks)} chunk(s) of up to {_PIPELINE_CHUNK_SIZE}.")

    clients = {'prod': prod_client, 'test': test_client}
    all_merged: List[Dict[str, Any]] = []
    all_drata: List[Dict[str, Any]] = []

    for chunk_num, chunk in enumerate(chunks, 1):
        if len(chunks) > 1:
            print(f"\n--- Chunk {chunk_num}/{len(chunks)} ({len(chunk)} users) ---")

        chunk_names = [u.get("Netbios_Name0") or u.get("netbios_name0") for u in chunk]
        chunk_names = [n for n in chunk_names if n]

        # Step 2: pull devices scoped to this chunk
        print(f"  Pulling devices scoped to {len(chunk_names)} user machine names ...")
        devices = pull_table(
            prod_client, args.devices, args.warehouse_prod, "devices",
            filter_sql=_names_filter(chunk_names),
            timeout=args.timeout,
        )
        if not devices:
            if chunk_num == 1 and len(chunks) == 1:
                print("  [FAIL] No devices matched the user set -- verify Netbios_Name0 alignment between users and devices tables")
                sys.exit(1)
            print(f"  [WARN] Chunk {chunk_num}: no devices matched -- skipping.")
            continue

        before_prefix = len(devices)
        devices = [
            d for d in devices
            if (d.get('Netbios_Name0') or d.get('Name0') or '').upper().startswith(DEVICE_NAME_PREFIXES)
        ]
        dropped_prefix = before_prefix - len(devices)
        if dropped_prefix:
            print(f"  [FILTER] {dropped_prefix} device(s) dropped -- name does not start with {DEVICE_NAME_PREFIXES}.")
        if not devices:
            print(f"  [WARN] Chunk {chunk_num}: no devices after prefix filter -- skipping.")
            continue

        resource_ids = [rid for rid in (get_resource_id(r) for r in devices) if rid is not None]
        filter_map = {
            'resource_id': _ids_filter(resource_ids),
            'netbios_name': _names_filter(chunk_names),
        }

        # Step 3: pull secondary tables for this chunk
        pulled: Dict[str, Any] = {}
        for spec in TABLE_REGISTRY:
            table_path = os.getenv(spec.env_var, '').strip()
            if not table_path:
                if spec.required:
                    print(f"  [FAIL] {spec.env_var} is required but not set")
                    sys.exit(1)
                pulled[spec.label] = None
                continue
            wh = args.warehouse_test if spec.client_key == 'test' else args.warehouse_prod
            if spec.batched:
                data = pull_table_batched(
                    clients[spec.client_key], table_path, wh, spec.label,
                    ids=resource_ids, timeout=args.timeout,
                )
            else:
                data = pull_table(
                    clients[spec.client_key], table_path, wh, spec.label,
                    filter_sql=filter_map[spec.filter_type], timeout=args.timeout,
                )
            pulled[spec.label] = data
            if not data:
                print(f"  [WARN] {spec.label} returned 0 rows")

        # Step 4: merge (user-centric inner join)
        print(f"  Merging (user-centric) ...")
        merged_chunk, dropped = merge(
            devices,
            pulled['windows_update'],
            pulled['installed_software'],
            chunk,
            bitlocker=pulled.get('bitlocker'),
            screensaver=pulled.get('screensaver'),
            services=pulled.get('services'),
            network_adapter=pulled.get('network_adapter'),
        )
        print(f"  {len(merged_chunk)} records assembled.")
        if dropped:
            print(f"  [INFO] {dropped} device(s) had no matching user in this chunk.")

        # Step 5: transform to Drata MDM format
        drata_chunk = transform_all(merged_chunk)
        if args.test_mode:
            drata_chunk = apply_test_overrides(drata_chunk)
        if args.sandbox:
            drata_chunk = apply_sandbox_overrides(drata_chunk)

        all_merged.extend(merged_chunk)
        all_drata.extend(drata_chunk)

    if args.test_mode:
        print(f"\n  [TEST MODE] {len(all_drata)} records with all 5 monitoring fields forced to passing.")
    else:
        print(f"\n  {len(all_drata)} total record(s) transformed.")
    if args.sandbox:
        print(f"  [SANDBOX] personnelId domain rewritten to @sandbox.nationwide.com.")

    # Step 6: write output files
    write_json(all_merged, raw_path)
    print(f"\n[OK] Raw merged JSON  : {raw_path}")

    write_json(all_drata, drata_path)
    print(f"[OK] Drata MDM JSON   : {drata_path}")

    # Step 7: push to Drata API
    # Filter records with no usable personnelId before pushing -- sending null to Drata
    # guarantees a 400; skip locally and log so the run doesn't burn retries on bad data.
    valid_payload = [r for r in all_drata if (r.get('personnelId') or '').strip()]
    no_pid = len(all_drata) - len(valid_payload)
    if no_pid:
        print(f"  [WARN] {no_pid} record(s) skipped -- personnelId is null or empty (UPN missing in source).")

    if args.dry_run:
        print(f"\n[DRY RUN] Would push {len(valid_payload)} records to Drata (skipped).\n")
    elif not api_key or not connection_id:
        print("\n[SKIP] DRATA_API_KEY or DRATA_CONNECTION_ID not set -- skipping push.\n")
    else:
        from db.drata_client import DrataClient
        print(f"\nPushing {len(valid_payload)} records to Drata (parallel) ...")
        drata = DrataClient(api_key=api_key, connection_id=connection_id)
        result = drata.push_batch_parallel(valid_payload)
        if result['errors']:
            print(f"  [WARN] {len(result['errors'])} record(s) failed -- review output above.")
        else:
            print(f"  [OK] Pushed {result['pushed']}/{result['total']} records.\n")


if __name__ == "__main__":
    main()
