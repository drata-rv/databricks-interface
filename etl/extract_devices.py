#!/usr/bin/env python3
"""
Device ETL: pulls SCCM tables from Databricks, joins them on resource_id
and Netbios_Name0, and produces JSON output for Drata Custom Device Connection.

Users are the authoritative anchor: only devices with a matched user record
are included in the output. Devices without a matching user are counted and logged.

Table configuration:
  - Users are loaded first and anchor all downstream scope.
  - Devices are pulled scoped to user machine names (no LIMIT).
  - TABLE_REGISTRY defines secondary tables (installed_software batched, test workspace).
  - required=True entries exit if env var not set.
  - required=False entries are skipped (null) when env var is empty.
  - Adding a new SCCM table: uncomment one registry line, set the env var.

Usage:
    python etl/extract_devices.py
    python etl/extract_devices.py --dry-run   # full pipeline, skip Drata push
    python etl/extract_devices.py --debug     # print resolved env before running
"""

import argparse
import atexit
import collections
import json
import os
import sys
import time
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

sys.path.insert(0, str(Path(__file__).parent.parent))

from db.auth import get_client_for_env, load_env
from db import queries
from db.queries import rows_to_records
from db.secrets import get_secret
from db.transform import transform_all, apply_test_overrides, apply_sandbox_overrides, decode_security_center_state


def _apply_cli_env_overrides() -> None:
    """Apply any --env KEY=VALUE pairs from argv to os.environ before anything else in this
    process reads an env var (argparse defaults below, TABLE_REGISTRY lookups, get_secret()'s
    scope resolution).

    Databricks serverless python_wheel_task has no environment-variable injection mechanism
    at any level -- job, task, or environment spec (confirmed against Jobs API docs and the
    serverless compute limitations page, which names job/task parameters as the replacement,
    2026-08-04). Job parameters (argv) are the only supported channel into the running
    process. This lets one job parameter carry any of this codebase's existing DATABRICKS_*
    env vars (secret scope, warehouse IDs, table paths) straight through to the exact same
    os.getenv()-based resolution already used everywhere else in db/auth.py, db/secrets.py,
    and TABLE_REGISTRY -- no dedicated CLI flag or extra plumbing needed per variable, and no
    change needed anywhere else when a new DATABRICKS_* var is added later.
    """
    argv = sys.argv[1:]
    i = 0
    while i < len(argv):
        if argv[i] == "--env" and i + 1 < len(argv):
            key, sep, value = argv[i + 1].partition("=")
            if sep:
                os.environ[key] = value
            i += 2
        else:
            i += 1


load_env()
_apply_cli_env_overrides()


# ---------------------------------------------------------------------------
# Crash safety: a fatal error (e.g. a required table exhausting retries) calls
# sys.exit() from deep inside the per-chunk loop. Without this, every chunk's
# already-merged, already-transformed output -- potentially hours of Databricks
# query time on a large --full run -- is discarded, since output files are only
# written after the entire chunk loop finishes. atexit fires on sys.exit() (and
# on an uncaught exception) from anywhere in the call stack, so this recovers
# whatever was completed without touching the loop's own control flow.
# ---------------------------------------------------------------------------
_crash_state: Dict[str, Any] = {
    'flushed_normally': False,
    'all_merged': [],
    'all_drata': [],
    'raw_path': None,
    'drata_path': None,
    'chunks_completed': 0,
    'total_chunks': 0,
}


def _flush_partial_on_crash() -> None:
    if _crash_state['flushed_normally'] or not _crash_state['all_merged']:
        return
    print(f"\n[CRASH RECOVERY] Pipeline aborted after "
          f"{_crash_state['chunks_completed']}/{_crash_state['total_chunks']} chunk(s) completed.")
    print(f"  Flushing {len(_crash_state['all_merged'])} already-merged record(s) to disk before exiting ...")
    write_json(_crash_state['all_merged'], _crash_state['raw_path'])
    write_json(_crash_state['all_drata'], _crash_state['drata_path'])
    print(f"  [OK] Partial raw JSON  : {_crash_state['raw_path']}")
    print(f"  [OK] Partial Drata JSON: {_crash_state['drata_path']}")
    print(f"  NOTE: these records were NOT pushed to Drata. Chunks are independent and Drata")
    print(f"        upserts on externalId, so fixing the failure and re-running from the start")
    print(f"        is safe and will not create duplicate Drata records.")


atexit.register(_flush_partial_on_crash)


# ---------------------------------------------------------------------------
# Internal column stripping
# ---------------------------------------------------------------------------
STRIP_PREFIXES = ("__",)
_MAX_RETRIES = 3

# Devices whose Netbios_Name0 does not start with one of these prefixes are dropped
# before the merge. Add new prefixes here as needed (case-insensitive).
DEVICE_NAME_PREFIXES = ('NW', 'GI')
_RETRY_DELAYS = (5, 15)  # seconds before attempt 2 and attempt 3
_SW_BATCH_SIZE = 200
_PERSONNEL_CHECK_WORKERS = 5
_PIPELINE_CHUNK_SIZE = 500

# All six secondary tables confirmed (2026-08-14, direct counts against si_prod_catalog) to
# be raw, periodically re-ingested landing tables -- same __date/__hour/__ingest_ts/__row_hash
# architecture as t_sccm_r_system, which was 3,074,075 rows for only 60,836 distinct devices
# before the same latest-batch-only filter was applied there. Confirmed batch counts:
# windows_update 1,888,445 rows/41,024 devices/50 batches; installed_software 166,528,900
# rows/41,015 devices/50 batches; antivirus_product 1,169,805/31,075/21; firewall_product
# 1,159,233/31,059/21; bitlocker (encryptable_volume) 747,263/36,296/21; computer_system
# 794,779/39,350/21. Every one needs the identical latest-batch-only treatment or every
# device's rows multiply across every historical ingestion batch.
_RAW_LANDING_TABLES = {
    'windows_update', 'installed_software', 'antivirus_product',
    'firewall_product', 'bitlocker', 'computer_system', 'screensaver',
}


def _latest_batch_clause(table_path: str) -> str:
    """Restrict a raw landing table to its single most recent (__date, __hour) ingestion
    batch -- see _RAW_LANDING_TABLES. Without this, every device/record appears once per
    historical batch instead of once, total."""
    return (
        f" AND __date = (SELECT MAX(__date) FROM {table_path})"
        f" AND __hour = (SELECT MAX(__hour) FROM {table_path}"
        f" WHERE __date = (SELECT MAX(__date) FROM {table_path}))"
    )


def _iamdb_personnel_statement(table: str) -> str:
    """Latest-batch, one-row-per-employeenumber pull of active personnel.

    employeetype='E'/employeestatus='A' confirmed 2026-08-24 against Nationwide's own
    21,337 current-active-personnel figure (matched within 1.5%; other employeetype codes
    are contractors/non-employees/service accounts, not personnel this connection covers).
    iamdb is not one-row-per-person even within a single batch -- GROUP BY collapses
    duplicate same-batch rows per employeenumber; MAX is an arbitrary but deterministic
    tiebreaker when a person's rows disagree, the same approach validated in the audit
    that confirmed this table/filter.
    """
    return f"""
        SELECT employeenumber, MAX(mail) AS mail, MAX(cn) AS cn
        FROM {table}
        WHERE __date = (SELECT MAX(__date) FROM {table})
          AND __hour = (SELECT MAX(__hour) FROM {table}
                         WHERE __date = (SELECT MAX(__date) FROM {table}))
          AND employeetype = 'E' AND employeestatus = 'A'
          AND employeenumber IS NOT NULL AND employeenumber != ''
        GROUP BY employeenumber
    """


def _sccm_employee_bridge_statement(table: str) -> str:
    """Most recent user_name0/windows_nt_domain0 per employee_number.

    Bridges iamdb's authoritative personnel record to SCCM's device-login username --
    devices only carry user_name0/user_domain0, not employee_number, so this bridge is
    required to scope devices to the correct personnel set. ROW_NUMBER (not independent
    MAX per column) keeps user_name0 and windows_nt_domain0 paired from the same row,
    since a person could plausibly change domain and username together (account migration).
    """
    return f"""
        SELECT employee_number, user_name0, windows_nt_domain0 FROM (
            SELECT employee_number, user_name0, windows_nt_domain0,
                   ROW_NUMBER() OVER (
                       PARTITION BY employee_number
                       ORDER BY __date DESC, __hour DESC, __row_hash DESC
                   ) AS __rn
            FROM {table}
            WHERE employee_number IS NOT NULL AND employee_number != ''
        ) WHERE __rn = 1
    """


def _drata_secret_names(sandbox: bool) -> Tuple[str, str, str]:
    """Returns (api_key_secret, connection_id_secret, host_secret) for the given sandbox flag.

    sandbox=True keeps the original, pre-existing secret names unchanged -- nothing in
    Databricks needs to change for a sandbox run to keep working exactly as it always has.
    sandbox=False uses new, deliberately distinct '-prod' names that must be added to the
    secret scope before a real production push is possible -- there is no fallback to the
    sandbox credentials; get_secret() raises loudly on Databricks if a key is missing, which
    is the correct failure mode here (never silently reuse sandbox credentials for a prod run).
    host_secret is optional in both cases -- callers must pass required=False when fetching
    it, so a missing key falls back to None (DrataClient's own default base URL) instead of
    raising. Sandbox and prod are not known to need different hosts today, only different
    credentials.
    """
    if sandbox:
        return "drata-api-key", "drata-connection-id", "drata-host"
    return "drata-api-key-prod", "drata-connection-id-prod", "drata-host-prod"


def is_internal(col: str) -> bool:
    return any(col.startswith(p) for p in STRIP_PREFIXES)


# ---------------------------------------------------------------------------
# Table registry
# ---------------------------------------------------------------------------
TableSpec = collections.namedtuple(
    'TableSpec', ['label', 'env_var', 'client_key', 'filter_type', 'required', 'batched']
)
# client_key   : 'prod' | 'test'
# filter_type  : 'resource_id' | 'netbios_name'
# required     : True = env var must be set; False = skipped (None) when empty
# batched      : True = use pull_table_batched() (IN-clause chunked to _SW_BATCH_SIZE)

TABLE_REGISTRY = [
    TableSpec('windows_update',     'DATABRICKS_TABLE_WINDOWS_UPDATE',     'prod', 'resource_id', True,  False),
    TableSpec('installed_software', 'DATABRICKS_TABLE_INSTALLED_SOFTWARE', 'prod', 'resource_id', True,  True),
    # Antivirus/firewall product tables (WSC-style, confirmed 2026-06-24). Batched
    # defensively -- row counts unconfirmed, treat as potentially large like installed_software.
    # antivirus_product feeds antivirusEnabled (presence-based: any registered row = protected;
    # see _antivirus_from_securitycenter in transform.py). firewall_product is pulled/merged
    # only -- transform.py has no reference to it, so it cannot reach the Drata push payload.
    TableSpec('antivirus_product',  'DATABRICKS_TABLE_ANTIVIRUS',           'prod', 'resource_id', False, True),
    TableSpec('firewall_product',   'DATABRICKS_TABLE_FIREWALL',            'prod', 'resource_id', False, True),
    # Confirmed 2026-07-15 (SCCM Test Tables - Updated 7-15.xlsx): t_sccm_gs_encryptable_volume
    # feeds encryptionEnabled (protection_status0); t_sccm_gs_computer_system feeds a real
    # hardware model (model0), fixing the prior CPU-type-as-model bug.
    TableSpec('bitlocker',          'DATABRICKS_TABLE_BITLOCKER',           'prod', 'resource_id', False, False),
    TableSpec('computer_system',    'DATABRICKS_TABLE_COMPUTER_SYSTEM',     'prod', 'resource_id', False, False),
    # 2026-08-26: screensaver settings found under a differently-named table (there is no
    # dedicated "screensaver" table) -- t_sccm_gs_desktop, confirmed from SCCM Test
    # Tables-August.xlsx's "Desktop" sheet, carries screen_saver_active0/secure0/timeout0
    # (the Control Panel\Desktop registry values SCCM inventories under that class name).
    TableSpec('screensaver',        'DATABRICKS_TABLE_SCREENSAVER',        'prod', 'resource_id', False, False),
    # Uncomment when Nationwide confirms table names:
    # TableSpec('services',        'DATABRICKS_TABLE_SERVICES',        'prod', 'resource_id', False, False),
    # TableSpec('network_adapter', 'DATABRICKS_TABLE_NETWORK_ADAPTER', 'prod', 'resource_id', False, False),
]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def clean(record: Dict[str, Any]) -> Dict[str, Any]:
    """Strip internal pipeline columns from a record."""
    return {k: v for k, v in record.items() if not is_internal(k)}


def get_resource_id(record: Dict[str, Any]) -> Optional[int]:
    """Resolve resource_id regardless of column name casing."""
    for key in ("resource_id", "ResourceID", "RESOURCEID"):
        if key in record:
            val = record[key]
            try:
                return int(val) if val is not None else None
            except (ValueError, TypeError):
                return None
    return None


def _ids_filter(ids: List[int], column: str = "resource_id") -> str:
    """Build a SQL IN filter for integer IDs. Returns '1=0' if the list is empty."""
    if not ids:
        return "1=0"
    return f"{column} IN ({', '.join(str(int(i)) for i in ids)})"


def _names_filter(names: List[str], column: str = "Netbios_Name0") -> str:
    """Build a SQL IN filter for string names with backslash and single-quote escaping.

    Backslash must be escaped first -- it's the escape metacharacter in Spark SQL string
    literals, so a name ending in a single backslash would otherwise escape the literal's
    closing quote and corrupt the rest of the generated IN-list.
    """
    if not names:
        return "1=0"
    escaped = ", ".join(
        "'" + n.replace("\\", "\\\\").replace("'", "''") + "'" for n in names
    )
    return f"{column} IN ({escaped})"


def _user_device_filter(usernames: List[str]) -> str:
    """SQL filter for devices scoped to a set of usernames, excluding non-compliant device types.

    All column references are from t_sccm_r_system. If the configured devices table
    lacks any of these columns, the query will error -- remove that clause.
    IS NULL fallbacks include devices where the field was not backfilled.
    """
    if not usernames:
        return "1=0"
    name_filter = _names_filter(usernames, column="user_name0")
    return (
        f"{name_filter}"
        f" AND (is_assigned_to_user0 = TRUE OR is_assigned_to_user0 IS NULL)"
        f" AND (is_virtual_machine0 = FALSE OR is_virtual_machine0 IS NULL)"
        f" AND (decommissioned0 IS NULL OR decommissioned0 = 0)"
        f" AND (obsolete0 IS NULL OR obsolete0 = 0)"
        f" AND (active0 IS NULL OR active0 != 0)"
        f" AND (operating_system_name_and0 NOT LIKE '%Server%' OR operating_system_name_and0 IS NULL)"
    )


def pull_table(
    client: Any,
    table: str,
    warehouse_id: str,
    label: str,
    limit: Optional[int] = None,
    filter_sql: Optional[str] = None,
    timeout: int = 300,
    required: bool = True,
    statement: Optional[str] = None,
) -> Optional[List[Dict[str, Any]]]:
    """Pull a table and return cleaned records.

    required=True (default): exits the process after exhausting retries -- used for
    tables the pipeline cannot proceed without (users, devices, windows_update, ...).
    required=False: logs a warning and returns None instead of exiting, so a transient
    failure on a best-effort table (e.g. antivirus_product) cannot abort a multi-hour run.
    statement: full SQL body to run verbatim instead of building SELECT * FROM table
    [WHERE filter_sql] [LIMIT limit] -- for pulls needing aggregation/window-function
    dedup beyond a simple WHERE filter (e.g. iamdb's latest-batch, one-row-per-person pull).
    """
    if statement is None:
        parts = [f"SELECT * FROM {table}"]
        if filter_sql:
            parts.append(f"WHERE {filter_sql}")
        if limit is not None:
            parts.append(f"LIMIT {limit}")
        statement = " ".join(parts)
    print(f"  Pulling {label} ({table}) ...")
    last_error: Optional[Exception] = None
    for attempt in range(1, _MAX_RETRIES + 1):
        try:
            result = queries.run_sql(
                client,
                statement=statement,
                warehouse_id=warehouse_id,
                timeout_seconds=timeout,
            )
            records = rows_to_records(result["columns"], result["rows"])
            print(f"  {len(records)} rows retrieved.")
            return [clean(r) for r in records]
        except Exception as e:
            last_error = e
            if attempt < _MAX_RETRIES:
                wait = _RETRY_DELAYS[attempt - 1]
                print(f"  [RETRY {attempt}/{_MAX_RETRIES}] {label} failed, retrying in {wait}s ...")
                time.sleep(wait)
    raw = str(last_error)
    short = raw.split(". Config:")[0].split(". Env:")[0].strip()
    if not required:
        print(f"  [WARN] {label} unavailable (all {_MAX_RETRIES} attempts failed) -- treating as absent for this run.")
        print(f"         Table     : {table}")
        print(f"         Error     : {short}")
        return None
    print(f"  [FAIL] {label} (all {_MAX_RETRIES} attempts failed)")
    print(f"         Table     : {table}")
    print(f"         Warehouse : {warehouse_id}")
    print(f"         Error     : {short}")
    sys.exit(1)


def pull_table_batched(
    client: Any,
    table: str,
    warehouse_id: str,
    label: str,
    ids: List[int],
    id_column: str = "resource_id",
    timeout: int = 300,
    required: bool = True,
    extra_filter: str = "",
) -> Optional[List[Dict[str, Any]]]:
    """Pull a large table by chunking the IN-clause into batches of _SW_BATCH_SIZE.

    When required=False, a batch that exhausts retries aborts just this table for this
    chunk (returns None) instead of exiting the whole process. extra_filter is appended
    (already including its own leading " AND ...") to every batch's filter -- see
    _latest_batch_clause() / _RAW_LANDING_TABLES.
    """
    all_records: List[Dict[str, Any]] = []
    total_batches = (len(ids) + _SW_BATCH_SIZE - 1) // _SW_BATCH_SIZE
    for i in range(0, len(ids), _SW_BATCH_SIZE):
        batch = ids[i : i + _SW_BATCH_SIZE]
        batch_num = i // _SW_BATCH_SIZE + 1
        records = pull_table(
            client, table, warehouse_id,
            f"{label} (batch {batch_num}/{total_batches})",
            filter_sql=_ids_filter(batch, column=id_column) + extra_filter,
            timeout=timeout,
            required=required,
        )
        if records is None:
            return None
        all_records.extend(records)
    return all_records


def merge(
    devices: List[Dict[str, Any]],
    windows_update: List[Dict[str, Any]],
    installed_software: List[Dict[str, Any]],
    users: List[Dict[str, Any]],
    bitlocker: Optional[List[Dict[str, Any]]] = None,
    screensaver: Optional[List[Dict[str, Any]]] = None,
    services: Optional[List[Dict[str, Any]]] = None,
    network_adapter: Optional[List[Dict[str, Any]]] = None,
    antivirus_product: Optional[List[Dict[str, Any]]] = None,
    firewall_product: Optional[List[Dict[str, Any]]] = None,
    computer_system: Optional[List[Dict[str, Any]]] = None,
) -> Tuple[List[Dict[str, Any]], int]:
    """
    Inner join: users are the anchor. Only devices with a matched user are included.

    Returns (records, dropped_count) where dropped_count is the number of devices
    that had no matching user entry.
    """
    # Build dual device indexes to support both join strategies:
    #   Netbios join  -- xlsx path (users have Netbios_Name0), inherently 1:1 (pre-joined)
    #   Username join -- Databricks path (users have user_name0 + windows_nt_domain0);
    #                    list-valued because one user can have more than one device
    #                    (e.g. desktop + laptop both last-logged-in by the same account) --
    #                    a plain-dict assignment here would silently drop all but the last.
    device_by_netbios: Dict[str, Dict[str, Any]] = {}
    device_by_username: Dict[Tuple[str, str], List[Dict[str, Any]]] = {}
    for dev in devices:
        netbios = dev.get('netbios_name0') or dev.get('name0') or dev.get('Netbios_Name0') or dev.get('Name0')
        if netbios:
            # Normalized (strip+lower) so a casing/whitespace difference between the xlsx
            # users source and the SCCM devices table can't silently defeat this join.
            device_by_netbios[netbios.strip().lower()] = dev
        uname = (dev.get('user_name0') or dev.get('User_Name0') or '').strip().lower()
        udomain = (dev.get('user_domain0') or dev.get('User_Domain0') or '').strip().lower()
        if uname:
            device_by_username.setdefault((uname, udomain), []).append(dev)

    # Index resource_id-keyed tables
    wu_index: Dict[int, Dict[str, Any]] = {}
    for row in windows_update:
        rid = get_resource_id(row)
        if rid is not None:
            wu_index[rid] = {k: v for k, v in row.items() if k not in ("resource_id", "ResourceID")}

    sw_index: Dict[int, List[Dict[str, Any]]] = {}
    for row in installed_software:
        rid = get_resource_id(row)
        if rid is not None:
            entry = {k: v for k, v in row.items() if k not in ("resource_id", "ResourceID")}
            sw_index.setdefault(rid, []).append(entry)

    # Optional tables -- build index only when table was pulled
    bitlocker_index: Optional[Dict[int, Dict[str, Any]]] = None
    if bitlocker is not None:
        bitlocker_index = {}
        for row in bitlocker:
            rid = get_resource_id(row)
            if rid is not None:
                bitlocker_index[rid] = {k: v for k, v in row.items() if k not in ("resource_id", "ResourceID")}

    screensaver_index: Optional[Dict[int, Dict[str, Any]]] = None
    if screensaver is not None:
        screensaver_index = {}
        for row in screensaver:
            rid = get_resource_id(row)
            if rid is not None:
                screensaver_index[rid] = {k: v for k, v in row.items() if k not in ("resource_id", "ResourceID")}

    services_index: Optional[Dict[int, List[Dict[str, Any]]]] = None
    if services is not None:
        services_index = {}
        for row in services:
            rid = get_resource_id(row)
            if rid is not None:
                entry = {k: v for k, v in row.items() if k not in ("resource_id", "ResourceID")}
                services_index.setdefault(rid, []).append(entry)

    network_adapter_index: Optional[Dict[int, Dict[str, Any]]] = None
    if network_adapter is not None:
        network_adapter_index = {}
        for row in network_adapter:
            rid = get_resource_id(row)
            if rid is not None:
                network_adapter_index[rid] = {k: v for k, v in row.items() if k not in ("resource_id", "ResourceID")}

    # antivirus_product / firewall_product: list-indexed like installed_software -- a device
    # can have more than one registered product row (e.g. a stale Defender entry alongside
    # an active third-party product).
    antivirus_index: Optional[Dict[int, List[Dict[str, Any]]]] = None
    if antivirus_product is not None:
        antivirus_index = {}
        for row in antivirus_product:
            rid = get_resource_id(row)
            if rid is not None:
                entry = {k: v for k, v in row.items() if k not in ("resource_id", "ResourceID")}
                antivirus_index.setdefault(rid, []).append(entry)

    firewall_index: Optional[Dict[int, List[Dict[str, Any]]]] = None
    if firewall_product is not None:
        firewall_index = {}
        for row in firewall_product:
            rid = get_resource_id(row)
            if rid is not None:
                entry = {k: v for k, v in row.items() if k not in ("resource_id", "ResourceID")}
                firewall_index.setdefault(rid, []).append(entry)

    computer_system_index: Optional[Dict[int, Dict[str, Any]]] = None
    if computer_system is not None:
        computer_system_index = {}
        for row in computer_system:
            rid = get_resource_id(row)
            if rid is not None:
                computer_system_index[rid] = {k: v for k, v in row.items() if k not in ("resource_id", "ResourceID")}

    # User-centric iteration: users anchor the output set.
    # Try Netbios join first (xlsx path), fall back to username+domain join (Databricks path).
    # A user can match more than one device on the Databricks path -- each produces its own
    # output record (same personnelId, different device), matching a real one-employee-many-
    # devices relationship instead of silently keeping only one.
    matched_device_ids: set = set()
    output: List[Dict[str, Any]] = []
    for row in users:
        netbios = row.get('Netbios_Name0') or row.get('netbios_name0')
        netbios_key = netbios.strip().lower() if netbios else None
        if netbios_key and netbios_key in device_by_netbios:
            matched_devices = [device_by_netbios[netbios_key]]
        else:
            uname = (row.get('user_name0') or '').strip().lower()
            udomain = (row.get('windows_nt_domain0') or '').strip().lower()
            matched_devices = device_by_username.get((uname, udomain), []) if uname else []
        for device in matched_devices:
            rid = get_resource_id(device)
            device_key = rid if rid is not None else (
                device.get('Netbios_Name0') or device.get('Name0') or device.get('netbios_name0')
            )
            matched_device_ids.add(device_key)
            if rid is None:
                label = device.get('Netbios_Name0') or device.get('Name0') or device.get('netbios_name0') or '?'
                print(f"  [WARN] merge: device {label!r} has no parseable resource_id -- software/WU data will be empty.")
            user_fields = {k: v for k, v in row.items() if k not in ('Netbios_Name0', 'netbios_name0')}
            device_fields = {k: v for k, v in device.items() if k not in ("resource_id", "ResourceID", "ResourceType")}
            output.append({
                "resource_id": rid,
                "device": device_fields,
                "windows_update": wu_index.get(rid, {}),
                "installed_software": sw_index.get(rid, []),
                "user": user_fields,
                "bitlocker": bitlocker_index.get(rid) if bitlocker_index is not None else None,
                "screensaver": screensaver_index.get(rid) if screensaver_index is not None else None,
                "services": services_index.get(rid, []) if services_index is not None else None,
                "network_adapter": network_adapter_index.get(rid) if network_adapter_index is not None else None,
                "antivirus_product": antivirus_index.get(rid, []) if antivirus_index is not None else None,
                "firewall_product": firewall_index.get(rid, []) if firewall_index is not None else None,
                "computer_system": computer_system_index.get(rid) if computer_system_index is not None else None,
            })

    dropped = len(devices) - len(matched_device_ids)
    return output, dropped


def write_json(payload: List[Dict[str, Any]], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, default=str)


_LOCAL_USERS_FILE = "SCCM Employees with Devices - Sandbox.xlsx"


def load_users_from_xlsx(path: str, netbios_filter: Optional[set] = None) -> List[Dict[str, Any]]:
    """Load user records from a local xlsx file. When netbios_filter is None, all rows are returned."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = {k: v for k, v in zip(headers, row) if k is not None}
        if netbios_filter is None or record.get('Netbios_Name0') in netbios_filter:
            records.append(record)
    wb.close()
    return records


def default_output_paths(test_mode: bool = False) -> Tuple[Path, Path]:
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    tag = "_test" if test_mode else ""
    return (
        Path("output") / f"devices_{timestamp}{tag}_raw.json",
        Path("output") / f"devices_{timestamp}{tag}_drata.json",
    )


# ---------------------------------------------------------------------------
# Args
# ---------------------------------------------------------------------------

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Pull SCCM tables from Databricks, merge by user, transform to Drata MDM format."
    )
    parser.add_argument(
        "--env",
        action="append",
        default=[],
        metavar="KEY=VALUE",
        help=(
            "Set an environment variable before any other config is resolved (repeatable). "
            "Already applied from sys.argv by _apply_cli_env_overrides() before this parser "
            "built its defaults -- registered here only so --help documents it and argparse "
            "doesn't reject it as unrecognized. This is how a Databricks job parameter "
            "carries a DATABRICKS_* env var into a serverless python_wheel_task, which has "
            "no environment-variable injection mechanism of its own."
        ),
    )
    parser.add_argument(
        "--devices",
        metavar="CATALOG.SCHEMA.TABLE",
        default=os.getenv("DATABRICKS_TABLE_DEVICES", ""),
        help="Fully qualified path to the main device table. Uses DATABRICKS_TABLE_DEVICES.",
    )
    parser.add_argument(
        "--devices-client",
        choices=["prod", "test"],
        default=os.getenv("DATABRICKS_DEVICES_CLIENT", "prod"),
        help=(
            "Which workspace connection to use for the devices table (default: prod, matching "
            "the real production data source). Set to 'test' for an end-to-end test run against "
            "a devices-equivalent table in the test workspace's catalog, so a first validation "
            "run needs no prod credentials at all. Uses DATABRICKS_DEVICES_CLIENT."
        ),
    )
    parser.add_argument(
        "--warehouse-prod",
        metavar="WAREHOUSE_ID",
        default=os.getenv("DATABRICKS_WAREHOUSE_ID", ""),
        help="Warehouse ID for the prod devices table. Uses DATABRICKS_WAREHOUSE_ID.",
    )
    parser.add_argument(
        "--warehouse-test",
        metavar="WAREHOUSE_ID",
        default=os.getenv("DATABRICKS_WAREHOUSE_ID_TEST", ""),
        help="Warehouse ID for the test catalog tables. Uses DATABRICKS_WAREHOUSE_ID_TEST.",
    )
    parser.add_argument(
        "--host-prod",
        metavar="URL",
        default=os.getenv("DATABRICKS_HOST_PROD", ""),
        help="Prod workspace URL. Uses DATABRICKS_HOST_PROD.",
    )
    parser.add_argument(
        "--host-test",
        metavar="URL",
        default=os.getenv("DATABRICKS_HOST_TEST", ""),
        help="Test workspace URL. Uses DATABRICKS_HOST_TEST.",
    )
    parser.add_argument(
        "--token-prod",
        metavar="TOKEN",
        default=os.getenv("DATABRICKS_TOKEN_PROD", "") or os.getenv("DATABRICKS_TOKEN", ""),
        help="Token for the prod workspace. Uses DATABRICKS_TOKEN_PROD, falls back to DATABRICKS_TOKEN.",
    )
    parser.add_argument(
        "--token-test",
        metavar="TOKEN",
        default=os.getenv("DATABRICKS_TOKEN_TEST", "") or os.getenv("DATABRICKS_TOKEN", ""),
        help="Token for the test workspace. Uses DATABRICKS_TOKEN_TEST, falls back to DATABRICKS_TOKEN.",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=int(os.getenv("DATABRICKS_LIMIT", "1000")),
        help="Max users to process per run (default: 1000). Devices and secondary tables are scoped to those users.",
    )
    parser.add_argument(
        "--timeout",
        type=int,
        default=int(os.getenv("DATABRICKS_QUERY_TIMEOUT", "300")),
        help="Per-query timeout in seconds (default: 300). Uses DATABRICKS_QUERY_TIMEOUT.",
    )
    parser.add_argument(
        "--output-raw",
        metavar="FILE",
        default="",
        help="Path for the raw merged JSON. Defaults to output/devices_<timestamp>_raw.json.",
    )
    parser.add_argument(
        "--output-drata",
        metavar="FILE",
        default="",
        help="Path for the Drata-formatted JSON. Defaults to output/devices_<timestamp>_drata.json.",
    )
    parser.add_argument(
        "--local-users",
        action="store_true",
        default=False,
        help=(
            f"Load the users table from the local xlsx file ({_LOCAL_USERS_FILE}) "
            "instead of pulling from Databricks. Bypasses DATABRICKS_TABLE_USERS and "
            "DATABRICKS_TABLE_IAMDB. Records are scoped to the machine names returned by "
            "the devices pull."
        ),
    )
    # test-mode/sandbox/dry-run/full accept an optional explicit value (nargs='?') rather than
    # plain store_true -- a Databricks job parameter is passed as "--test-mode" "{{...}}" (two
    # separate tokens), which store_true can't represent since it takes no value at all. Bare
    # `--test-mode` (no value, local CLI usage) still works via const='true'.
    parser.add_argument(
        "--test-mode",
        nargs='?', const='true', default='false',
        help=(
            "Push real identities to Drata with all 5 monitoring fields forced to a passing "
            "state. Uses real personnelId/alias/externalId so records land on actual users. "
            "Combine with --limit to control how many are processed. "
            "Intended for verifying the Drata connection end-to-end."
        ),
    )
    parser.add_argument(
        "--sandbox",
        nargs='?', const='true', default='false',
        help="Replace @nationwide.com with @sandbox.nationwide.com in personnelId before pushing.",
    )
    parser.add_argument(
        "--dry-run",
        nargs='?', const='true', default='false',
        help="Run the full pipeline but skip the Drata API push. Output files are still written.",
    )
    parser.add_argument(
        "--debug",
        action="store_true",
        default=False,
        help="Print full resolved config and env var sources before running.",
    )
    parser.add_argument(
        "--full",
        nargs='?', const='true', default='false',
        help=(
            "Process all users -- bypasses --limit and runs the full dataset using the "
            "chunked pipeline. Intended for production sync of 25,000+ users."
        ),
    )
    args = parser.parse_args()
    for flag in ('test_mode', 'sandbox', 'dry_run', 'full'):
        setattr(args, flag, str(getattr(args, flag)).strip().lower() in ('true', '1', 'yes'))
    return args


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    args = parse_args()

    # Validate required non-credential config upfront. Host/token are deliberately NOT
    # checked here -- they're resolved by get_client_for_env() via db.secrets.get_secret(),
    # which may source them from a Databricks secret scope rather than a plain env var when
    # running inside a Databricks Job (where args.host_prod's argparse default would be empty
    # even though the real value is available). get_client_for_env()/WorkspaceClient() raise
    # their own clear error if host/token are genuinely unset in either environment.
    required_values = [
        ("--devices (or DATABRICKS_TABLE_DEVICES)", args.devices),
        ("--warehouse-test (or DATABRICKS_WAREHOUSE_ID_TEST)", args.warehouse_test),
    ]
    # warehouse-prod is only required when devices actually comes from the prod workspace --
    # --devices-client test (a prod-free end-to-end test run) has no other use for it.
    if args.devices_client == "prod":
        required_values.append(("--warehouse-prod (or DATABRICKS_WAREHOUSE_ID)", args.warehouse_prod))
    missing = [name for name, val in required_values if not val.strip()]

    if missing:
        print("Error: the following required values are not set:")
        for m in missing:
            print(f"  {m}")
        sys.exit(1)

    # 2026-08-17: test_client construction is conditional again -- secret scopes are
    # workspace-specific, and dbx_irg_dev_comp_sp only exists in the TEST workspace. This
    # job now also runs on compute hosted in the PROD workspace (-t prod), where that scope
    # does not exist at all -- an unconditional get_client_for_env("test") failed there with
    # "Secret does not exist with scope: dbx_irg_dev_comp_sp" before any prod logic ran, on
    # a real prod deploy 2026-08-17. prod_client stays unconditional -- users and every
    # current TABLE_REGISTRY entry use client_key='prod' regardless of --devices-client.
    # test_client is only actually needed for a --devices-client test override (which only
    # makes sense run against the test/dev target/workspace anyway).
    prod_client = get_client_for_env("prod")
    test_client = get_client_for_env("test") if args.devices_client == "test" else prod_client
    default_raw, default_drata = default_output_paths(test_mode=args.test_mode)
    raw_path = Path(args.output_raw) if args.output_raw else default_raw
    drata_path = Path(args.output_drata) if args.output_drata else default_drata
    # Preserve the conventional _drata.json -> _rejected.json naming when present, but never
    # let rejected_path collapse to the same path as drata_path for a --output-drata value
    # that doesn't contain that exact substring -- that would silently clobber the real
    # output file when the rejected-records write happens later.
    if '_drata.json' in drata_path.name:
        rejected_name = drata_path.name.replace('_drata.json', '_rejected.json')
    else:
        rejected_name = drata_path.stem + '_rejected' + drata_path.suffix
    rejected_path = drata_path.parent / rejected_name
    _crash_state['raw_path'] = raw_path
    _crash_state['drata_path'] = drata_path

    print(f"\nProd workspace   : {args.host_prod}")
    print(f"Test workspace   : {args.host_test}")
    print(f"Warehouse (prod) : {args.warehouse_prod}")
    print(f"Warehouse (test) : {args.warehouse_test}")
    print(f"Limit (users)    : {args.limit} users")
    print(f"Query timeout    : {args.timeout}s per table")
    print(f"Output (raw)     : {raw_path}")
    print(f"Output (drata)   : {drata_path}")
    if args.local_users:
        print(f"Users source     : LOCAL FILE ({_LOCAL_USERS_FILE})")
    if args.test_mode:
        print(f"Mode             : TEST MODE (all 5 monitoring fields forced passing)")
    if args.full:
        print(f"Mode             : FULL SYNC (all users, --limit bypassed)")
    if args.dry_run:
        print(f"Mode             : DRY RUN (Drata push skipped)")

    if args.debug:
        databrickscfg = Path.home() / ".databrickscfg"
        print(f"\n-- DEBUG --")
        print(f"DATABRICKS_HOST_PROD             : {os.getenv('DATABRICKS_HOST_PROD', '(not set)')}")
        print(f"DATABRICKS_HOST_TEST             : {os.getenv('DATABRICKS_HOST_TEST', '(not set)')}")
        print(f"DATABRICKS_TOKEN_PROD            : {'(set)' if os.getenv('DATABRICKS_TOKEN_PROD') or os.getenv('DATABRICKS_TOKEN') else '(not set)'}")
        print(f"DATABRICKS_TOKEN_TEST            : {'(set)' if os.getenv('DATABRICKS_TOKEN_TEST') or os.getenv('DATABRICKS_TOKEN') else '(not set)'}")
        print(f"DATABRICKS_WAREHOUSE_ID          : {os.getenv('DATABRICKS_WAREHOUSE_ID', '(not set)')}")
        print(f"DATABRICKS_WAREHOUSE_ID_TEST     : {os.getenv('DATABRICKS_WAREHOUSE_ID_TEST', '(not set)')}")
        print(f"DATABRICKS_TABLE_DEVICES         : {os.getenv('DATABRICKS_TABLE_DEVICES', '(not set)')}")
        print(f"DATABRICKS_TABLE_USERS           : {os.getenv('DATABRICKS_TABLE_USERS', '(not set)')}")
        print(f"DATABRICKS_TABLE_IAMDB           : {os.getenv('DATABRICKS_TABLE_IAMDB', '(not set)')}")
        for spec in TABLE_REGISTRY:
            val = os.getenv(spec.env_var, '(not set)')
            req = 'required' if spec.required else 'optional'
            print(f"{spec.env_var:<40}: {val}  [{req}]")
        print(f"DATABRICKS_QUERY_TIMEOUT         : {os.getenv('DATABRICKS_QUERY_TIMEOUT', '(not set, using 300)')}")
        _dbg_api_key_secret, _dbg_conn_id_secret, _dbg_host_secret = _drata_secret_names(args.sandbox)
        print(f"Drata tenant (sandbox={args.sandbox}) secrets: {_dbg_api_key_secret}, {_dbg_conn_id_secret}, {_dbg_host_secret}")
        print(f"{_dbg_api_key_secret.upper().replace('-', '_'):<34}: {'(set)' if os.getenv(_dbg_api_key_secret.upper().replace('-', '_')) else '(not set)'}")
        print(f"{_dbg_conn_id_secret.upper().replace('-', '_'):<34}: {os.getenv(_dbg_conn_id_secret.upper().replace('-', '_'), '(not set)')}")
        print(f"~/.databrickscfg exists          : {databrickscfg.exists()}")
        if args.local_users:
            xlsx_exists = Path(_LOCAL_USERS_FILE).exists()
            print(f"LOCAL_USERS_FILE                 : {_LOCAL_USERS_FILE}  ({'found' if xlsx_exists else 'NOT FOUND'})")
        print(f"-- END DEBUG --\n")
    else:
        print()

    # 2026-08-24: which Drata tenant we talk to (sandbox vs prod) is selected here, purely
    # by which secret names get resolved -- see _drata_secret_names(). Nothing else in this
    # file needs to know which tenant is active; api_key/connection_id/drata_host are used
    # identically either way, for both the personnel-status check below and the final push.
    api_key_secret, connection_id_secret, host_secret = _drata_secret_names(args.sandbox)
    api_key = (get_secret(api_key_secret) or "").strip()
    connection_id = (get_secret(connection_id_secret) or "").strip()
    drata_host = (get_secret(host_secret, required=False) or "").strip() or None

    # Step 1: load users first (anchor for all downstream scope)
    if args.local_users:
        if not Path(_LOCAL_USERS_FILE).exists():
            print(f"  [FAIL] Local users file not found: {_LOCAL_USERS_FILE}")
            sys.exit(1)
        print(f"  [LOCAL] Loading users from {_LOCAL_USERS_FILE} ...")
        all_users = load_users_from_xlsx(_LOCAL_USERS_FILE, netbios_filter=None)
        print(f"  {len(all_users)} users loaded.")
    else:
        iamdb_table = os.getenv("DATABRICKS_TABLE_IAMDB", "").strip()
        sccm_users_table = os.getenv("DATABRICKS_TABLE_USERS", "").strip()
        if not iamdb_table or not sccm_users_table:
            print("  [FAIL] DATABRICKS_TABLE_IAMDB and DATABRICKS_TABLE_USERS are both required (or use --local-users)")
            sys.exit(1)

        # 2026-08-24: personnel anchor switched from t_sccm_r_user's own identity fields to
        # t_iamdb_userdata (Nationwide's authoritative IAM source), joined to SCCM via
        # employee_number for whichever active personnel actually have a device to report on.
        # This fixes personnelId being sourced from SCCM's own copy of the user's email,
        # which is what caused the @Nationwide.com casing bug (iamdb's mail is the
        # authoritative value, not a copy). The Drata personnel-status check below is
        # unrelated and still runs -- iamdb answers "is this person active per Nationwide's
        # HR system," Drata's own check answers "does Drata itself recognize this person,"
        # a different question that can disagree with iamdb.
        personnel = pull_table(
            prod_client, iamdb_table, args.warehouse_prod, "iamdb personnel",
            statement=_iamdb_personnel_statement(iamdb_table), timeout=args.timeout,
        )
        print(f"  {len(personnel)} active personnel (employeetype=E, employeestatus=A).")

        bridge = pull_table(
            prod_client, sccm_users_table, args.warehouse_prod, "sccm employee bridge",
            statement=_sccm_employee_bridge_statement(sccm_users_table), timeout=args.timeout,
        )
        bridge_by_empnum = {b['employee_number']: b for b in bridge}
        print(f"  {len(bridge_by_empnum)} distinct employee_number(s) with an SCCM device-login record.")

        all_users = []
        no_device_record = 0
        for p in personnel:
            b = bridge_by_empnum.get(p['employeenumber'])
            if b is None:
                no_device_record += 1
                continue
            all_users.append({
                'user_name0': b['user_name0'],
                'windows_nt_domain0': b['windows_nt_domain0'],
                'mail': p['mail'],
                'employeenumber': p['employeenumber'],
            })
        print(f"  {len(all_users)} active personnel matched to an SCCM device-login record "
              f"({no_device_record} active personnel have no SCCM device record -- no device to report on).")

    if args.full:
        print(f"  Full sync: processing all {len(all_users)} users.")
    elif args.limit and len(all_users) > args.limit:
        print(f"  Limiting to {args.limit} users (of {len(all_users)} total).")
        all_users = all_users[:args.limit]

    # Step 1b: verify each candidate against Drata's own personnel record. Restored
    # 2026-08-24 -- Drata is the actual system of record for whether a pushed device can
    # correctly link to a real personnel entry there; iamdb's employeestatus is Nationwide's
    # HR truth, a different question that can disagree with Drata's own roster (confirmed:
    # Drata's Personnel view reports a different total than iamdb's active-employee count).
    if api_key:
        from db.drata_client import DrataClient
        from concurrent.futures import ThreadPoolExecutor, as_completed as _as_completed
        drata_pre = DrataClient(api_key=api_key, connection_id=connection_id, base_url=drata_host)
        _active = {'CURRENT_EMPLOYEE', 'CURRENT_CONTRACTOR'}
        before = len(all_users)
        sandbox_flag = args.sandbox

        def _check_one(u):
            try:
                email = (
                    u.get('mail')                    # Databricks path: iamdb (authoritative)
                    or u.get('User_Princiipal_Name0')   # xlsx: double-i typo
                    or u.get('User_Principal_Name0') # xlsx: correct spelling
                    or u.get('user_principal_name0') # legacy fallback
                    or ''
                ).lower()
                if not email or '@' not in email:
                    return u, None
                lookup = email
                if sandbox_flag and '@nationwide.com' in email:
                    lookup = email.replace('@nationwide.com', '@sandbox.nationwide.com')
                return u, drata_pre.get_person_status(lookup)
            except Exception:
                return u, '__error__'

        filtered = []
        check_errors = 0
        print(f"Checking {before} user(s) against Drata personnel status ({_PERSONNEL_CHECK_WORKERS} workers) ...")
        with ThreadPoolExecutor(max_workers=_PERSONNEL_CHECK_WORKERS) as pool:
            futures = {pool.submit(_check_one, u): u for u in all_users}
            for i, fut in enumerate(_as_completed(futures), 1):
                u, status = fut.result()
                if status == '__error__':
                    check_errors += 1
                elif status in _active:
                    filtered.append(u)
                if i % 500 == 0:
                    print(f"  ... {i}/{before} checked, {len(filtered)} active ...")
        all_users = filtered
        skipped = before - len(all_users) - check_errors
        print(f"  Drata personnel filter: {len(all_users)} confirmed / {skipped} excluded (not found/former in Drata) / {check_errors} API errors (excluded).")
    else:
        print(f"  [WARN] {api_key_secret} not set -- Drata personnel verification skipped, all users will be processed.")

    if not all_users:
        print("  [FAIL] No users remain after Drata personnel verification -- check DATABRICKS_TABLE_IAMDB/DATABRICKS_TABLE_USERS, the Drata credentials, or user data")
        sys.exit(1)

    # Steps 2-5: process in chunks to bound memory and Databricks query scope
    chunks = [all_users[i:i + _PIPELINE_CHUNK_SIZE]
              for i in range(0, len(all_users), _PIPELINE_CHUNK_SIZE)]
    print(f"\n{len(all_users)} user(s) in {len(chunks)} chunk(s) of up to {_PIPELINE_CHUNK_SIZE}.")
    _crash_state['total_chunks'] = len(chunks)

    clients = {'prod': prod_client, 'test': test_client}
    all_merged: List[Dict[str, Any]] = []
    all_drata: List[Dict[str, Any]] = []
    av_state_counts: Dict[Tuple[str, Any], int] = {}
    # Alias (not copy) -- _crash_state stays current via in-place .extend() below with no
    # further updates needed here.
    _crash_state['all_merged'] = all_merged
    _crash_state['all_drata'] = all_drata

    for chunk_num, chunk in enumerate(chunks, 1):
        if len(chunks) > 1:
            print(f"\n--- Chunk {chunk_num}/{len(chunks)} ({len(chunk)} users) ---")

        # Step 2: pull devices scoped to this chunk.
        # xlsx path: join key is Netbios_Name0 (pre-joined in the spreadsheet).
        # Databricks path: join key is user_name0 + domain; server/shared exclusion applied at SQL level.
        if args.local_users:
            chunk_names = [u.get("Netbios_Name0") or u.get("netbios_name0") for u in chunk]
            chunk_names = [n for n in chunk_names if n]
            device_filter = _names_filter(chunk_names)
            print(f"  Pulling devices scoped to {len(chunk_names)} machine names (Netbios_Name0) ...")
        else:
            chunk_names = [u.get("user_name0") for u in chunk]
            chunk_names = [n for n in chunk_names if n]
            device_filter = _user_device_filter(chunk_names)
            print(f"  Pulling devices scoped to {len(chunk_names)} usernames (user_name0, excluding servers/shared) ...")

        # t_sccm_r_system is a raw, periodically re-ingested landing table, not a
        # deduplicated device master -- confirmed 2026-08-07 via a real run: 3,074,075
        # rows for only 60,836 distinct resource_id (~56 rows/device, uniform across
        # every device checked), clustered on __date/__hour. Every past ingestion batch
        # is still in the table. Without this filter the same device appears dozens of
        # times, all sharing the same fallback externalId once serial_number/aad_device_id
        # are absent (common on this table) -- this is what rejected ~50,000 of ~50,483
        # transformed records as externalId collisions on the 2026-08-07 full run.
        # Restricting to the single latest (__date, __hour) batch brings it back to one
        # row per device. 2026-08-18: applied unconditionally instead of gated on
        # devices_client == "test" -- prod's devices_table now also points at
        # t_sccm_r_system (Terry Hardaway confirmed the "v_..._temp" table prod
        # previously used was not the intended source), the same raw multi-batch table
        # as test, so it needs the same fix.
        device_filter += _latest_batch_clause(args.devices)

        devices_client = prod_client if args.devices_client == "prod" else test_client
        devices_warehouse = args.warehouse_prod if args.devices_client == "prod" else args.warehouse_test
        devices = pull_table(
            devices_client, args.devices, devices_warehouse, "devices",
            filter_sql=device_filter,
            timeout=args.timeout,
        )
        if not devices:
            if chunk_num == 1 and len(chunks) == 1:
                print("  [FAIL] No devices matched the user set -- verify join key alignment (Netbios_Name0 for xlsx, user_name0 for Databricks)")
                sys.exit(1)
            print(f"  [WARN] Chunk {chunk_num}: no devices matched -- skipping.")
            continue

        before_prefix = len(devices)
        devices = [
            d for d in devices
            if (d.get('Netbios_Name0') or d.get('Name0') or d.get('netbios_name0') or d.get('name0') or '').upper().startswith(DEVICE_NAME_PREFIXES)
        ]
        dropped_prefix = before_prefix - len(devices)
        if dropped_prefix:
            print(f"  [FILTER] {dropped_prefix} device(s) dropped -- name does not start with {DEVICE_NAME_PREFIXES}.")
        if not devices:
            print(f"  [WARN] Chunk {chunk_num}: no devices after prefix filter -- skipping.")
            continue

        resource_ids = [rid for rid in (get_resource_id(r) for r in devices) if rid is not None]
        # Derived from the pulled devices themselves, not chunk_names -- chunk_names holds
        # usernames (not Netbios names) in the default Databricks mode, so a table registered
        # with filter_type='netbios_name' would otherwise compare Netbios_Name0 against
        # username values and match nothing.
        device_names = [
            d.get('Netbios_Name0') or d.get('Name0') or d.get('netbios_name0') or d.get('name0')
            for d in devices
        ]
        device_names = [n for n in device_names if n]
        filter_map = {
            'resource_id': _ids_filter(resource_ids),
            'netbios_name': _names_filter(device_names),
        }

        # Step 3: pull secondary tables for this chunk
        pulled: Dict[str, Any] = {}
        for spec in TABLE_REGISTRY:
            table_path = os.getenv(spec.env_var, '').strip()
            if not table_path:
                if spec.required:
                    print(f"  [FAIL] {spec.env_var} is required but not set")
                    sys.exit(1)
                pulled[spec.label] = None
                continue
            wh = args.warehouse_test if spec.client_key == 'test' else args.warehouse_prod
            latest_batch = _latest_batch_clause(table_path) if spec.label in _RAW_LANDING_TABLES else ""
            if spec.batched:
                data = pull_table_batched(
                    clients[spec.client_key], table_path, wh, spec.label,
                    ids=resource_ids, timeout=args.timeout, required=spec.required,
                    extra_filter=latest_batch,
                )
            else:
                data = pull_table(
                    clients[spec.client_key], table_path, wh, spec.label,
                    filter_sql=filter_map[spec.filter_type] + latest_batch, timeout=args.timeout, required=spec.required,
                )
            pulled[spec.label] = data
            if data is not None and not data:
                print(f"  [WARN] {spec.label} returned 0 rows")

        # Diagnostic only -- tally product_state0 combos seen so far; antivirusEnabled
        # in this run's push is NOT derived from this data (see [DIAGNOSTIC] summary below).
        if pulled.get('antivirus_product'):
            for row in pulled['antivirus_product']:
                key = (row.get('display_name0') or '(no name)', row.get('product_state0'))
                av_state_counts[key] = av_state_counts.get(key, 0) + 1

        # Step 4: merge (user-centric inner join)
        print(f"  Merging (user-centric) ...")
        merged_chunk, dropped = merge(
            devices,
            pulled['windows_update'],
            pulled['installed_software'],
            chunk,
            bitlocker=pulled.get('bitlocker'),
            screensaver=pulled.get('screensaver'),
            services=pulled.get('services'),
            network_adapter=pulled.get('network_adapter'),
            antivirus_product=pulled.get('antivirus_product'),
            firewall_product=pulled.get('firewall_product'),
            computer_system=pulled.get('computer_system'),
        )
        print(f"  {len(merged_chunk)} records assembled.")
        if dropped:
            print(f"  [INFO] {dropped} device(s) had no matching user in this chunk.")

        # Step 5: transform to Drata MDM format
        drata_chunk = transform_all(merged_chunk)
        if args.test_mode:
            drata_chunk = apply_test_overrides(drata_chunk)
        if args.sandbox:
            drata_chunk = apply_sandbox_overrides(drata_chunk)

        all_merged.extend(merged_chunk)
        all_drata.extend(drata_chunk)
        _crash_state['chunks_completed'] = chunk_num

    if not all_merged:
        print("  [FAIL] No records produced across all chunks.")
        print("         Possible causes: Netbios_Name0 mismatch between users and devices tables,")
        print("         all users excluded by personnel filter, or all devices dropped by prefix filter.")
        sys.exit(1)

    if args.test_mode:
        print(f"\n  [TEST MODE] {len(all_drata)} records with all 5 monitoring fields forced to passing.")
    else:
        print(f"\n  {len(all_drata)} total record(s) transformed.")
    if args.sandbox:
        print(f"  [SANDBOX] personnelId domain rewritten to @sandbox.nationwide.com.")

    if av_state_counts:
        total_rows = sum(av_state_counts.values())
        print(f"\n[DIAGNOSTIC] antivirus_product product_state0 distribution "
              f"({len(av_state_counts)} distinct combo(s) across {total_rows} row(s)):")
        print(f"  antivirusEnabled treats ANY row in this table as a confirmed AV signal "
              f"(presence-based). The decode column below is informational only -- "
              f"it is not used to determine pass/fail.")
        ranked = sorted(av_state_counts.items(), key=lambda kv: -kv[1])
        for (name, state), count in ranked[:30]:
            decoded = decode_security_center_state(state)
            tag = {True: 'ENABLED', False: 'DISABLED', None: 'UNKNOWN'}[decoded]
            print(f"    {count:>6}x  display_name0={name!r:<40} product_state0={state!r:<12} decode={tag}")
        if len(ranked) > 30:
            print(f"    ... and {len(ranked) - 30} more combo(s) -- see {raw_path} for full antivirus_product data")

    # 2026-08-26: auoptions0 (autoUpdateEnabled's source column, t_sccm_gs_windowsupdate)
    # was never confirmed against a real DESCRIBE -- prod run same day showed the check
    # failing for effectively everyone. _ci_get() in db/transform.py now matches the key
    # case-insensitively, but if the real column is named something else entirely this
    # still won't resolve it -- print the actual explanation distribution and raw column
    # names seen so a wrong guess is visible immediately instead of needing another round trip.
    if all_drata:
        au_explanation_counts: Dict[Optional[str], int] = {}
        for rec in all_drata:
            key = rec.get('autoUpdateExplanation')
            au_explanation_counts[key] = au_explanation_counts.get(key, 0) + 1
        wu_key_samples: set = set()
        for rec in all_merged:
            wu = rec.get('windows_update')
            if wu:
                wu_key_samples.update(wu.keys())
        print(f"\n[DIAGNOSTIC] autoUpdateEnabled explanation distribution across {len(all_drata)} record(s):")
        for explanation, count in sorted(au_explanation_counts.items(), key=lambda kv: -kv[1]):
            print(f"    {count:>6}x  {explanation!r}")
        print(f"  windows_update raw column(s) seen this run: {sorted(wu_key_samples)}")

    # Step 6: write output files
    write_json(all_merged, raw_path)
    print(f"\n[OK] Raw merged JSON  : {raw_path}")

    write_json(all_drata, drata_path)
    print(f"[OK] Drata MDM JSON   : {drata_path}")
    _crash_state['flushed_normally'] = True

    # Step 7: push to Drata API
    # Each filter below removes records that cannot be safely pushed and logs them to
    # the rejected file so data quality gaps are visible without polluting the push run.
    rejected: List[Dict[str, Any]] = []

    # No personnelId -- cannot link device to a person in Drata
    no_pid_records = [r for r in all_drata if not (r.get('personnelId') or '').strip()]
    for r in no_pid_records:
        rejected.append({**r, 'rejection_reason': 'missing_personnelId'})
    valid_payload = [r for r in all_drata if (r.get('personnelId') or '').strip()]
    if no_pid_records:
        print(f"  [WARN] {len(no_pid_records)} record(s) excluded -- personnelId null or empty (UPN missing in source).")

    # No appList -- SCCM has no software inventory for this device; pushing an empty list
    # is a data quality issue, not a real device state.
    no_applist_records = [r for r in valid_payload if not r.get('appList')]
    for r in no_applist_records:
        rejected.append({**r, 'rejection_reason': 'empty_applist'})
    valid_payload = [r for r in valid_payload if r.get('appList')]
    if no_applist_records:
        print(f"  [WARN] {len(no_applist_records)} record(s) excluded -- appList empty (no software inventory in SCCM).")

    # No externalId -- Drata uses externalId as a upsert key; without it a device cannot
    # be matched on subsequent runs and creates duplicate orphaned records.
    no_extid_records = [r for r in valid_payload if not r.get('externalId') or r.get('externalId') == 'None']
    for r in no_extid_records:
        rejected.append({**r, 'rejection_reason': 'missing_externalId'})
    valid_payload = [r for r in valid_payload if r.get('externalId') and r.get('externalId') != 'None']
    if no_extid_records:
        print(f"  [WARN] {len(no_extid_records)} record(s) excluded -- externalId null (serial number missing in SCCM).")

    # Duplicate externalId -- would silently collide on the Drata upsert key, with one
    # device's compliance record overwriting another's. Defense in depth on top of the
    # placeholder-serial-number rejection in transform.py: catches any other source of
    # a shared value (data entry error, cloned image, etc.) that isn't a known placeholder.
    extid_counts: Dict[Any, int] = {}
    for r in valid_payload:
        extid_counts[r.get('externalId')] = extid_counts.get(r.get('externalId'), 0) + 1
    dup_ids = {eid for eid, count in extid_counts.items() if count > 1}
    if dup_ids:
        dup_records = [r for r in valid_payload if r.get('externalId') in dup_ids]
        for r in dup_records:
            rejected.append({**r, 'rejection_reason': 'duplicate_externalId'})
        valid_payload = [r for r in valid_payload if r.get('externalId') not in dup_ids]
        print(f"  [WARN] {len(dup_records)} record(s) excluded -- externalId collides with another "
              f"record in this run ({len(dup_ids)} distinct colliding value(s)).")

    if rejected:
        write_json(rejected, rejected_path)
        print(f"  [WARN] {len(rejected)} total record(s) rejected -- see {rejected_path}")

    if args.dry_run:
        print(f"\n[DRY RUN] Would push {len(valid_payload)} records to Drata (skipped).\n")
    elif not api_key or not connection_id:
        print(f"\n[SKIP] {api_key_secret} or {connection_id_secret} not set -- skipping push.\n")
    else:
        from db.drata_client import DrataClient
        print(f"\nPushing {len(valid_payload)} records to Drata (parallel) ...")
        drata = DrataClient(api_key=api_key, connection_id=connection_id, base_url=drata_host)
        result = drata.push_batch_parallel(valid_payload)
        if result['errors']:
            # 2026-08-19: reverted the 2026-08-07 change that sys.exit(1)'d here on any
            # push error. This pipeline is meant to be resilient -- log errors and run to
            # completion, not abort a multi-hour full run over a handful of bad records
            # (real-world SCCM data always has some: missing fields, transient API blips,
            # etc.). Per-record failures are already fully visible here and in the rejected
            # file; a human reviews those separately from whether the job itself crashed.
            print(f"  [FAIL] {len(result['errors'])}/{result['total']} record(s) failed to push.")
            print("  Failed records:")
            for err in result['errors'][:20]:
                print(f"    personnelId={err.get('personnelId')}  alias={err.get('alias')}  error={err.get('error')}")
            if len(result['errors']) > 20:
                print(f"    ... and {len(result['errors']) - 20} more (see {drata_path} for full payload)")
            print(f"  Pushed {result['pushed']}/{result['total']} records; {len(result['errors'])} failed.\n")
        else:
            print(f"  [OK] Pushed {result['pushed']}/{result['total']} records.\n")


if __name__ == "__main__":
    main()
